from Controllers.Clientes._cliente import *
from Models.pagamentoModel import *
from Models.arquivoModel import *
from openpyxl import load_workbook

# CLASSE DO LANÇAMENTO DE PROVISIONAMENTO
class ProvisionamentoCliente(Cliente):

    def __init__(self):
        super().__init__()

    # GERENCIA O LANÇAMENTO DAS ATAS
    def varrer(self, db, login, senha):
        # url_remoto, url_arquivos = self.format_paths(db)

        self.driver.get(self.pagina_inicial)
        try:
            if not self.login(login, senha, db):
                return False
        except:
            tb = traceback.format_exc()
            print(tb)
            print('Erro no Login')
            time.sleep(60)
            return False

        planilha = load_workbook(filename="C:\\temp\\provisionamento.xlsx")
        provs = planilha['Plan1']
        i = 0
        for row in provs.iter_rows():
            i += 1
            if i == 1:
                continue

            print(row[2].value)
            if row[2].value == "OK":
                continue

            self.driver.get(self.pagina_busca)
            self.busca_processo(row[0].value.strip())

            if self.lanca_provisao_trabalhista(row[1].value):
                provs["C"+str(i)] = "OK"
            else:
                if self.lanca_provisao_trabalhista(row[1].value):
                    provs["C"+str(i)] = "OK"

            planilha.save(filename="C:\\temp\\provisionamento.xlsx")



        return True

    def lanca_provisao_trabalhista(self,valor):
        self.abre_fecha_aba('PEDIDO(S)')

        # for tr in trs:
        titulos = ['VERBAS RESCISORIAS', 'FGTS']
        for t in titulos:
            achei = False
            pg1 = self.driver.find_element_by_xpath('//*[@id="fDetalhar:scrollResultadoddtbPedidoidx1"]')
            if pg1:
                pg1.click()

            while True:
                trs = self.driver.find_elements_by_xpath('//*[@id="fDetalhar:dtbPedido"]/tbody/tr')
                for idx, tr in enumerate(trs):
                    # aguarda_presenca_elemento(self.driver, '//*[@id="fDetalhar:dtbPedido"]/tbody/tr['+str(idx+1)+']/td[1]')
                    titulo = self.driver.find_element_by_xpath('//*[@id="fDetalhar:dtbPedido"]/tbody/tr['+str(idx+1)+']/td[1]').text

                    if titulo == 'TOTAL':
                        achei = True
                        break

                    if not titulo == t:
                        continue

                    achei = True
                    btn = self.driver.find_element_by_xpath('//*[@id="fDetalhar:dtbPedido"]/tbody/tr[' + str(idx + 1) + ']/td[10]/a')

                    valor_longo_prazo = self.driver.find_element_by_xpath('//*[@id="fDetalhar:dtbPedido"]/tbody/tr[' + str(idx + 1) + ']/td[2]').text.strip()

                    if valor_longo_prazo == '2.222,00':
                        return True

                    btn.click()
                    self.abre_e_aguarda_modal()

                    self.seleciona_option('1', 'fPedido:codTipoReavaliacaoLongoProvavel', tipo='ID', select_by='value')
                    self.preenche_campo('2.222,00', 'fPedido:valorLongoProvavelT', tipo='ID')
                    file_upload = self.driver.find_element_by_id('fPedido:anexo')
                    url_arq = trata_path("C:\\temp\\INGRESSO VIKSTAR.pdf")
                    file_upload.send_keys(url_arq)
                    self.driver.find_element_by_xpath('//*[@id="fPedido:_idJsp24"]').click()

                    self.driver.find_element_by_id('fPedido:_idJsp21').click()
                    self.fecha_modal()

                if achei:
                    break

                pg2 = self.driver.find_element_by_xpath('//*[@id="fDetalhar:scrollResultadoddtbPedidonext"]')
                if pg2:
                    pg2.click()

            if achei:
                break

        self.abre_fecha_aba('PEDIDO(S)')
        return False